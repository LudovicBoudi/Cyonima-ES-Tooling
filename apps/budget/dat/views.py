import csv
from io import BytesIO
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import DAT, DATLine
from apps.budget.providers.models import Provider


@login_required
def dat_list(request):
    dats = DAT.objects.all().select_related('provider').prefetch_related('lines')
    year = request.GET.get('year', '')
    if year:
        dats = dats.filter(year=int(year))
    years = DAT.objects.values_list('year', flat=True).distinct().order_by('-year')
    providers = Provider.objects.all().order_by('company_name')
    return render(request, 'budget/dat_list.html', {
        'dats': dats,
        'years': years,
        'providers': providers,
        'selected_year': year,
        'STATUS_CHOICES': DAT.STATUS_CHOICES,
    })


@login_required
def dat_detail(request, pk):
    dat = get_object_or_404(DAT.objects.prefetch_related('lines').select_related('provider'), pk=pk)
    return render(request, 'budget/dat_detail.html', {'dat': dat})


def _safe_int(val, default=1):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _safe_float(val, default=0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


@login_required
def dat_create(request):
    if request.method == 'POST':
        provider_id = request.POST.get('provider')
        if not provider_id:
            messages.error(request, "Un fournisseur est requis.")
            providers = Provider.objects.all()
            return render(request, 'budget/dat_form.html', {'providers': providers})
        provider = get_object_or_404(Provider, pk=provider_id)
        dat = DAT.objects.create(
            provider=provider,
            description=request.POST.get('description', ''),
            provider_contact_name=request.POST.get('contact_name', ''),
            provider_contact_email=request.POST.get('contact_email', ''),
            provider_contact_phone=request.POST.get('contact_phone', ''),
        )
        products = request.POST.getlist('product[]')
        refs = request.POST.getlist('reference[]')
        quantities = request.POST.getlist('quantity[]')
        units = request.POST.getlist('unit[]')
        prices = request.POST.getlist('unit_price[]')
        btypes = request.POST.getlist('budget_type[]')
        bcats = request.POST.getlist('budget_category[]')
        for i, product in enumerate(products):
            if product.strip():
                DATLine.objects.create(
                    dat=dat, product=product,
                    reference=refs[i] if i < len(refs) else '',
                    unit=units[i] if i < len(units) else 'U',
                    quantity=_safe_int(quantities[i] if i < len(quantities) else '', 1),
                    unit_price=_safe_float(prices[i] if i < len(prices) else '', 0),
                    budget_type=btypes[i] if i < len(btypes) else 'investment',
                    budget_category=bcats[i] if i < len(bcats) else 'pc',
                )
        messages.success(request, f'DAT {dat.display_id()} créée.')
        return redirect('dat_detail', pk=dat.pk)
    providers = Provider.objects.all()
    return render(request, 'budget/dat_form.html', {'providers': providers})


@login_required
def dat_update(request, pk):
    dat = get_object_or_404(DAT.objects.prefetch_related('lines'), pk=pk)
    if dat.status in ('validated', 'rejected'):
        messages.error(request, "Impossible de modifier une DAT validée ou refusée.")
        return redirect('dat_detail', pk=pk)
    if request.method == 'POST':
        provider_id = request.POST.get('provider')
        dat.provider = get_object_or_404(Provider, pk=provider_id)
        dat.description = request.POST.get('description', '')
        dat.provider_contact_name = request.POST.get('contact_name', '')
        dat.provider_contact_email = request.POST.get('contact_email', '')
        dat.provider_contact_phone = request.POST.get('contact_phone', '')
        dat.save()
        dat.lines.all().delete()
        products = request.POST.getlist('product[]')
        refs = request.POST.getlist('reference[]')
        quantities = request.POST.getlist('quantity[]')
        units = request.POST.getlist('unit[]')
        prices = request.POST.getlist('unit_price[]')
        btypes = request.POST.getlist('budget_type[]')
        bcats = request.POST.getlist('budget_category[]')
        for i, product in enumerate(products):
            if product.strip():
                DATLine.objects.create(
                    dat=dat, product=product,
                    reference=refs[i] if i < len(refs) else '',
                    unit=units[i] if i < len(units) else 'U',
                    quantity=_safe_int(quantities[i] if i < len(quantities) else '', 1),
                    unit_price=_safe_float(prices[i] if i < len(prices) else '', 0),
                    budget_type=btypes[i] if i < len(btypes) else 'investment',
                    budget_category=bcats[i] if i < len(bcats) else 'pc',
                )
        messages.success(request, f'DAT {dat.display_id()} modifiée.')
        return redirect('dat_detail', pk=dat.pk)
    providers = Provider.objects.all()
    return render(request, 'budget/dat_form.html', {'providers': providers, 'dat': dat})


@login_required
def dat_delete(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if dat.status in ('validated', 'rejected'):
        messages.error(request, "Impossible de supprimer une DAT validée ou refusée.")
        return redirect('dat_detail', pk=pk)
    if request.method == 'POST':
        dat.delete()
        messages.success(request, 'DAT supprimée.')
        return redirect('dat_list')
    return render(request, 'budget/dat_confirm_delete.html', {'dat': dat})


# --- Workflow actions ---

@login_required
def dat_submit(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if request.method == 'POST':
        if dat.status != 'draft':
            messages.error(request, "Seul un brouillon peut être soumis.")
        else:
            dat.status = 'submitted'
            dat.save()
            messages.success(request, f'DAT {dat.display_id()} soumise pour validation.')
    return redirect('dat_detail', pk=pk)


@login_required
def dat_validate(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if request.method == 'POST':
        if dat.status != 'submitted':
            messages.error(request, "Seule une DAT soumise peut être validée.")
        else:
            dat.status = 'validated'
            dat.save()
            messages.success(request, f'DAT {dat.display_id()} validée.')
    return redirect('dat_detail', pk=pk)


@login_required
def dat_reject(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if request.method == 'POST':
        if dat.status != 'submitted':
            messages.error(request, "Seule une DAT soumise peut être refusée.")
        else:
            dat.status = 'rejected'
            dat.save()
            messages.success(request, f'DAT {dat.display_id()} refusée.')
    return redirect('dat_detail', pk=pk)


@login_required
def dat_draft(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if request.method == 'POST':
        if dat.status not in ('submitted', 'rejected'):
            messages.error(request, "Seule une DAT soumise ou refusée peut repasser en brouillon.")
        else:
            dat.status = 'draft'
            dat.save()
            messages.success(request, f'DAT {dat.display_id()} repassée en brouillon.')
    return redirect('dat_detail', pk=pk)


@login_required
def dat_duplicate(request, pk):
    dat = get_object_or_404(DAT, pk=pk)
    if request.method == 'POST':
        new_dat = dat.duplicate()
        messages.success(request, f'DAT dupliquée. Nouvel ID : {new_dat.display_id()}')
        return redirect('dat_detail', pk=new_dat.pk)
    return redirect('dat_detail', pk=pk)

# --- Exports ---

def _safe_csv(val):
    s = str(val)
    if s and s[0] in ('=', '+', '-', '@'):
        return "'" + s
    return s


@login_required
def dat_export_csv(request, pk):
    dat = get_object_or_404(DAT.objects.prefetch_related('lines'), pk=pk)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="DAT_{dat.display_id()}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['ID', 'Date', 'Description', 'Fournisseur', 'Contact', 'Email', 'Téléphone',
                     'Produit', 'Référence', 'Qté', 'Prix unitaire', 'Total ligne',
                     'Budget', 'Catégorie', 'Total DAT'])
    lines = list(dat.lines.all())
    if lines:
        first = True
        for line in lines:
            if first:
                writer.writerow([
                    dat.display_id(), dat.created_date, dat.description,
                    dat.provider.company_name, dat.provider_contact_name,
                    dat.provider_contact_email, dat.provider_contact_phone,
                    line.product, line.reference, line.quantity,
                    str(line.unit_price).replace('.', ','),
                    str(line.global_price).replace('.', ','),
                    line.get_budget_type_display(), line.get_budget_category_display(),
                    str(dat.total_cost()).replace('.', ','),
                ])
                first = False
            else:
                writer.writerow([
                    '', '', '', '', '', '', '',
                    line.product, line.reference, line.quantity,
                    str(line.unit_price).replace('.', ','),
                    str(line.global_price).replace('.', ','),
                    line.get_budget_type_display(), line.get_budget_category_display(), '',
                ])
    else:
        writer.writerow([
            dat.display_id(), dat.created_date, dat.description,
            dat.provider.company_name, dat.provider_contact_name,
            dat.provider_contact_email, dat.provider_contact_phone,
            '', '', '', '', '', '', '', str(dat.total_cost()).replace('.', ','),
        ])
    return response


@login_required
def dat_export_xlsx(request, pk):
    dat = get_object_or_404(DAT.objects.prefetch_related('lines'), pk=pk)
    wb = Workbook()
    ws = wb.active
    ws.title = f'DAT {dat.display_id()}'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A3A6B', end_color='1A3A6B', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.cell(row=1, column=1, value='DAT').font = Font(bold=True, size=14, color='1A3A6B')
    ws.merge_cells('A1:G1')

    info = [
        ('ID', dat.display_id()), ('Date', str(dat.created_date)),
        ('Fournisseur', dat.provider.company_name), ('Contact', dat.provider_contact_name),
        ('Email', dat.provider_contact_email), ('Téléphone', dat.provider_contact_phone),
        ('Description', dat.description),
    ]
    for i, (label, val) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

    data_start = 3 + len(info) + 1
    headers = ['Produit', 'Référence', 'Qté', 'Prix unitaire (€)', 'Total (€)', 'Budget', 'Catégorie']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, line in enumerate(dat.lines.all(), start=data_start + 1):
        vals = [
            line.product, line.reference, line.quantity,
            float(line.unit_price), float(line.global_price),
            line.get_budget_type_display(), line.get_budget_category_display(),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            if col in (3, 4, 5):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00' if col > 3 else '0'

    total_row = data_start + 1 + dat.lines.count()
    cell = ws.cell(row=total_row, column=1, value='Total')
    cell.font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    cell = ws.cell(row=total_row, column=5, value=float(dat.total_cost()))
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal='right')

    for col_letter, width in [('A', 25), ('B', 25), ('C', 10), ('D', 15), ('E', 15), ('F', 20), ('G', 20)]:
        ws.column_dimensions[col_letter].width = width

    filename = f'DAT_{dat.display_id().replace("-", "_")}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def dat_export_pdf(request, pk):
    from weasyprint import HTML
    dat = get_object_or_404(DAT.objects.prefetch_related('lines').select_related('provider'), pk=pk)
    html = render_to_string('budget/dat_pdf.html', {'dat': dat})
    pdf = HTML(string=html).write_pdf()
    filename = f'DAT_{dat.display_id().replace("-", "_")}.pdf'
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
